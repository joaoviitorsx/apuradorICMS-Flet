from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill 
from sqlalchemy.orm import Session
from src.Models.c170cloneModel import C170Clone 
from src.Models._0150Model import Registro0150

COLUNAS = [
    'id', 'empresa_id', 'id_c100', 'ind_oper', 'filial', 'periodo', 'reg', 'cod_part',
    'nome', 'cnpj', 'num_doc', 'cod_item', 'chv_nfe', 'num_item', 'descr_compl', 'ncm', 'unid',
    'qtd', 'vl_item', 'vl_desc', 'cfop', 'cst', 'aliquota', 'resultado'
]

class ExportarPlanilhaRepository:
    def __init__(self, session: Session):
        self.session = session

    def buscarGalera(self, empresa_id: int, periodo: str):
        query = (
            self.session.query(
                C170Clone,
                Registro0150.nome,
                Registro0150.cnpj
            )
            .outerjoin(
                Registro0150,
                (Registro0150.cod_part == C170Clone.cod_part) &
                (Registro0150.empresa_id == C170Clone.empresa_id) &
                (Registro0150.periodo == C170Clone.periodo) &
                (Registro0150.is_active == True)
            )
            .filter(
                C170Clone.empresa_id == empresa_id,
                C170Clone.periodo == periodo,
                C170Clone.is_active == True
            )
        )
        return query

class ExportarPlanilhaService:
    def __init__(self, session: Session):
        self.session = session
        self.repository = ExportarPlanilhaRepository(session)

    def exportarC170Clone(self, empresa_id: int, periodo: str, caminho_saida: str) -> dict:
        print(f"[DEBUG] exportarC170Clone chamado: empresa_id={empresa_id}, periodo={periodo}, caminho_saida={caminho_saida}")
        try:
            query = self.repository.buscarGalera(empresa_id, periodo)
            total = query.count()
            
            print(f"[DEBUG] Total de registros ativos encontrados: {total}")
            if total == 0:
                return {"status": "vazio", "mensagem": "Nenhum registro ativo encontrado para exportação."}

            wb = Workbook()
            ws = wb.active
            ws.title = f"Exportação {periodo.replace('/', '-')}"

            ws.append(COLUNAS)
            
            for i, col in enumerate(COLUNAS, start=1):
                cell = ws.cell(row=1, column=i)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

            linhas_exportadas = 0

            for c170, nome, cnpj in query.yield_per(1000):
                linha = []
                for col in COLUNAS:
                    if col == "nome":
                        linha.append(nome or "")
                    elif col == "cnpj":
                        linha.append(cnpj or "")
                    else:
                        valor = getattr(c170, col, None)
                        linha.append(valor if valor is not None else "")
                
                ws.append(linha)
                linhas_exportadas += 1
                
                if linhas_exportadas % 1000 == 0:
                    print(f"[DEBUG] {linhas_exportadas} linhas exportadas...")

            print(f"[DEBUG] Total de linhas ativas exportadas: {linhas_exportadas}")

            for i, col in enumerate(COLUNAS, start=1):
                ws.column_dimensions[get_column_letter(i)].width = max(15, len(col) + 3)

            wb.save(caminho_saida)
            print(f"[DEBUG] Planilha salva em: {caminho_saida}")
            return {
                "status": "ok", 
                "mensagem": f"Planilha exportada com sucesso! {linhas_exportadas} registros ativos exportados.",
                "registros_exportados": linhas_exportadas,
                "caminho": caminho_saida
            }

        except Exception as e:
            print(f"[DEBUG] Erro ao exportar planilha: {e}")
            return {"status": "erro", "mensagem": f"Erro ao exportar planilha: {str(e)}"}