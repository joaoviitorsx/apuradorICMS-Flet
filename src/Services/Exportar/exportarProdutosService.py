import os
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill
from sqlalchemy.orm import Session
from src.Models.tributacaoModel import CadastroTributacao

class ExportarProdutosRepository:
    def __init__(self, session: Session):
        self.session = session

    def buscarProdutos(self, empresa_id: int):
        query = (
            self.session.query(CadastroTributacao)
            .filter(CadastroTributacao.empresa_id == empresa_id)
            .order_by(CadastroTributacao.codigo, CadastroTributacao.produto)
        )
        return query

class ExportarProdutosService:
    def __init__(self, session: Session):
        self.session = session
        self.repository = ExportarProdutosRepository(session)

    def exportarProdutos(self, empresa_id: int, caminho_saida: str) -> dict:
        print(f"[DEBUG] Iniciando exportação de produtos: empresa_id={empresa_id}")
        print(f"[DEBUG] Caminho de saída original: {caminho_saida}")
        
        if not caminho_saida.lower().endswith('.xlsx'):
            caminho_saida = caminho_saida + '.xlsx'
        
        print(f"[DEBUG] Caminho de saída final: {caminho_saida}")
        
        try:
            diretorio = os.path.dirname(caminho_saida)
            if diretorio and not os.path.exists(diretorio):
                os.makedirs(diretorio, exist_ok=True)
                print(f"[DEBUG] Diretório criado: {diretorio}")
            
            query = self.repository.buscarProdutos(empresa_id)
            total = query.count()
            print(f"[DEBUG] Total de produtos encontrados: {total}")
            
            if total == 0:
                return {
                    "status": "aviso", 
                    "mensagem": "Nenhum produto encontrado para exportação.",
                    "total_produtos": 0
                }

            wb = Workbook()
            ws = wb.active
            ws.title = "Produtos_Tributacao"

            cabecalho = ['Empresa ID', 'Código', 'Produto', 'NCM', 'Alíquota (%)', 'Categoria Fiscal']
            ws.append(cabecalho)
            
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            for i, _ in enumerate(cabecalho, start=1):
                cell = ws.cell(row=1, column=i)
                cell.font = header_font
                cell.fill = header_fill

            linhas_exportadas = 0
            
            for produto in query.yield_per(1000):
                try:
                    linha = [
                        produto.empresa_id,           # Empresa ID
                        produto.codigo or "",         # Código
                        produto.produto or "",        # Produto
                        produto.ncm or "",           # NCM
                        produto.aliquota or 0,       # Alíquota
                        produto.categoriaFiscal or "" # Categoria Fiscal
                    ]
                    ws.append(linha)
                    linhas_exportadas += 1
                    
                    if linhas_exportadas % 500 == 0:
                        print(f"[DEBUG] {linhas_exportadas} produtos exportados...")
                        
                except Exception as e:
                    print(f"[DEBUG] Erro ao processar produto ID {produto.id}: {e}")
                    continue

            print(f"[DEBUG] Total de produtos exportados: {linhas_exportadas}")

            colunas_largura = {
                1: 12,  # Empresa ID
                2: 15,  # Código
                3: 40,  # Produto
                4: 12,  # NCM
                5: 12,  # Alíquota
                6: 20   # Categoria Fiscal
            }
            
            for col_num, largura in colunas_largura.items():
                ws.column_dimensions[get_column_letter(col_num)].width = largura

            wb.save(caminho_saida)
            print(f"[DEBUG] Planilha salva com sucesso: {caminho_saida}")
            
            if os.path.exists(caminho_saida):
                tamanho_arquivo = os.path.getsize(caminho_saida)
                print(f"[DEBUG] Arquivo criado com sucesso! Tamanho: {tamanho_arquivo} bytes")
            else:
                print(f"[DEBUG] ERRO: Arquivo não foi criado!")
            
            return {
                "status": "ok", 
                "mensagem": f"✅ Exportação concluída!\n\n📊 {linhas_exportadas} produtos exportados\n📁 Arquivo salvo em: {caminho_saida}",
                "total_produtos": linhas_exportadas,
                "caminho_arquivo": caminho_saida
            }

        except FileNotFoundError:
            return {
                "status": "erro", 
                "mensagem": "❌ Caminho de destino não encontrado. Verifique se o diretório existe."
            }
        except PermissionError:
            return {
                "status": "erro", 
                "mensagem": "❌ Permissão negada. Verifique se o arquivo não está aberto em outro programa."
            }
        except Exception as e:
            print(f"[DEBUG] Erro inesperado ao exportar produtos: {e}")
            import traceback
            traceback.print_exc()
            return {
                "status": "erro", 
                "mensagem": f"❌ Erro inesperado ao exportar produtos: {str(e)}"
            }